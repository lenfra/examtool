# -*- coding: utf-8 -*-
from ExamClasses.StudentExamGrade import StudentExamGrade
from openpyxl import load_workbook
import re
from datetime import datetime


class ExamResultParser:
    def __init__(self, connector, settings, exam_db):
        self.cnx = connector
        self.Settings = settings
        self._exam_db = exam_db

        self._exam_id = None
        self._exam_date = None

        self._course_id = None
        self.result_started = False

        self._question_map = []
        self._student_order = 0

    def set_exam_id(self, exam_id):
        if self._exam_db.exam_exist(exam_id):
            self._exam_id = exam_id
            return True
        else:
            print("ExamID don't exist")
            return False

    def get_exam_id(self):
        return self._exam_id

    def set_exam_date(self, exam_date):
        if isinstance(exam_date, str):
            exam_date = datetime.strptime(exam_date, '%Y-%m-%d')

        assert (isinstance(exam_date, datetime))

        self._exam_date = exam_date

    def get_exam_date(self):
        return self.set_exam_date

    def set_course_id(self, _course_id):
        self._course_id = _course_id

    def get_course_id(self):
        return self._course_id

    def load_result(self, filename):
        _workbook = load_workbook(filename, True)
        worksheet = _workbook.get_sheet_by_name("Result")

        for row in worksheet.iter_rows():
            self._parse_row(row)

        return True

    def _parse_row(self, row):
        _course_id_re = re.compile('Course: (.+?)$')
        _exam_id_re = re.compile('Exam ID: (.+?)$')
        _exam_date_re = re.compile('Examdate: (.+?)$')
        _student_code_re = re.compile('Student Code:')
        _result_row = self.result_started
        student_result = None
        question_number = 0

        if _result_row:
            student_result = StudentExamGrade(exam_id=self._exam_id)
            student_result.set_connector(self.cnx)

        for cell in row:
            _cell_value = cell.value
            if _cell_value is not None:
                if not isinstance(_cell_value, str):
                    _cell_value = str(_cell_value)

                if _course_id_re.search(_cell_value):
                    self.set_course_id(_course_id_re.findall(_cell_value)[0])

                if _exam_id_re.search(_cell_value):
                    if not self.set_exam_id(_exam_id_re.findall(_cell_value)[0]):
                        return False

                if _exam_date_re.search(_cell_value):
                    self.set_exam_date(_exam_date_re.findall(_cell_value)[0])

                if _student_code_re.search(_cell_value):
                    self.result_started = True

                # Map column with question ID
                if not _result_row:
                    if self._course_id:
                        if re.search(self._course_id + 'q[0-9]*$', _cell_value):
                            question_id = re.findall(self._course_id + 'q[0-9]*$', _cell_value)[0]
                            self._question_map.append([cell.column, question_id])

                if _result_row:
                    if cell.column == 1:  # Student ID
                        student_result.set_student_id(_cell_value)
                        student_result.set_student_order(self._student_order)
                        self._student_order += 1
                    for _c in self._question_map:
                        if cell.column == _c[0]:
                            student_result.append_question_grade(_c[1], _cell_value, row[cell.column].value,
                                                                 question_number=question_number)
                            question_number += 1
        if student_result:
            if student_result.insert_into_database():
                return True

        return False
